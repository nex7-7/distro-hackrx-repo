"""
Document processing service for the RAG Application API.

This module handles Stage 1 of the RAG pipeline: document ingestion and preprocessing.
It downloads documents, classifies file types, and extracts text using various libraries
based on the file format.
"""

import asyncio
import hashlib
import mimetypes
import tempfile
import zipfile
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union
from urllib.parse import urlparse
import time
from datetime import datetime

import requests
import pymupdf4llm
import pdfplumber
from docx import Document as DocxDocument
from openpyxl import load_workbook
import pytesseract
from PIL import Image
import email
from email.policy import default

from config.settings import settings
from app.models.schemas import DocumentInfo
from app.utils.logger import get_logger
from app.utils.exceptions import (
    DocumentDownloadError,
    UnsupportedFileTypeError,
    DocumentExtractionError,
    create_error
)

logger = get_logger(__name__)


class DocumentProcessor:
    """
    Document processing service that handles file download, classification, and text extraction.
    
    This class implements Stage 1 of the RAG pipeline with support for multiple file formats
    and intelligent extraction strategies.
    """
    
    # Supported file types and their MIME types
    SUPPORTED_TYPES = {
        '.pdf': 'application/pdf',
        '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        '.txt': 'text/plain',
        '.eml': 'message/rfc822',
        '.png': 'image/png',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.zip': 'application/zip'
    }
    
    def __init__(self) -> None:
        """Initialize the document processor."""
        # Create timestamped download directory with proper error handling
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        project_root = Path(__file__).parent.parent.parent  # Navigate to project root
        
        try:
            self.downloads_dir = project_root / "downloads" / f"data_{timestamp}"
            self.downloads_dir.mkdir(parents=True, exist_ok=True, mode=0o777)
        except (PermissionError, OSError) as e:
            # Fall back to temp directory if downloads directory is not writable
            import tempfile
            temp_dir = Path(tempfile.gettempdir()) / "rag_downloads" / f"data_{timestamp}"
            temp_dir.mkdir(parents=True, exist_ok=True)
            self.downloads_dir = temp_dir
            logger.warning(f"Could not create downloads directory in {project_root / 'downloads'}: {e}. Using temporary directory: {self.downloads_dir}")
        
        logger.info(f"Downloads directory set to: {self.downloads_dir}")
        
        # Configure Tesseract path if specified
        if settings.tesseract_cmd:
            pytesseract.pytesseract.tesseract_cmd = settings.tesseract_cmd
            logger.info("Tesseract configured", cmd_path=settings.tesseract_cmd)
        else:
            # Try to detect Tesseract automatically
            try:
                # Test if tesseract is available in PATH
                version = pytesseract.get_tesseract_version()
                logger.info("Tesseract found in PATH", version=str(version))
            except pytesseract.TesseractNotFoundError:
                logger.warning("Tesseract not found in PATH. Image processing will fail unless TESSERACT_CMD is set.")
        
        logger.info("Document processor initialized", downloads_dir=str(self.downloads_dir))
    
    async def process_document(self, document_url: str) -> DocumentInfo:
        """
        Main entry point for document processing.
        
        Args:
            document_url: URL of the document to process
            
        Returns:
            DocumentInfo: Processed document information
            
        Raises:
            DocumentDownloadError: If document download fails
            UnsupportedFileTypeError: If file type is not supported
            DocumentExtractionError: If text extraction fails
        """
        start_time = time.time()
        logger.log_stage("Document Processing", "Starting", url=document_url)
        
        try:
            # Step 1: Download the document
            file_path, filename = await self._download_document(document_url)
            
            # Step 2: Classify file type
            file_type = self._classify_file_type(file_path, filename)
            
            # Step 3: Extract text content
            extracted_text, extraction_method, total_pages = await self._extract_text(
                file_path, file_type
            )
            
            # Step 4: Generate content hash
            content_hash = self._generate_content_hash(extracted_text)
            
            # Step 5: Extract first 5 pages for context
            first_five_pages = self._extract_first_pages(extracted_text, 5)
            
            # Create document info
            doc_info = DocumentInfo(
                url=document_url,
                filename=filename,
                file_type=file_type,
                content_hash=content_hash,
                total_pages=total_pages,
                extracted_text=extracted_text,
                first_five_pages=first_five_pages,
                extraction_method=extraction_method
            )
            
            duration = time.time() - start_time
            logger.log_performance(
                "Document Processing",
                duration,
                file_type=file_type,
                text_length=len(extracted_text),
                pages=total_pages
            )
            
            return doc_info
            
        except Exception as e:
            logger.error(f"Document processing failed: {str(e)}", url=document_url)
            raise
        # Note: Files are now stored permanently in downloads directory for data management
    
    async def _download_document(self, url: str) -> Tuple[Path, str]:
        """
        Download document from URL.
        
        Args:
            url: Document URL
            
        Returns:
            Tuple[Path, str]: File path and filename
            
        Raises:
            DocumentDownloadError: If download fails
        """
        try:
            logger.debug("Starting document download", url=url)
            
            # Parse URL to get filename
            parsed_url = urlparse(url)
            filename = Path(parsed_url.path).name or "document"
            
            # Download with streaming to handle large files
            response = requests.get(url, stream=True, timeout=30)
            response.raise_for_status()
            
            # Save to downloads directory
            file_path = self.downloads_dir / filename
            with open(file_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)
            
            file_size = file_path.stat().st_size
            logger.info("Document downloaded successfully", 
                       filename=filename, 
                       size_bytes=file_size)
            
            return file_path, filename
            
        except requests.RequestException as e:
            raise create_error(
                DocumentDownloadError,
                f"Failed to download document: {str(e)}",
                "DOWNLOAD_FAILED",
                url=url,
                error_type=type(e).__name__
            )
    
    def _classify_file_type(self, file_path: Path, filename: str) -> str:
        """
        Classify file type based on extension and MIME type.
        
        Args:
            file_path: Path to the file
            filename: Original filename
            
        Returns:
            str: File type (extension without dot)
            
        Raises:
            UnsupportedFileTypeError: If file type is not supported
        """
        # Get extension from filename
        file_extension = Path(filename).suffix.lower()
        
        # Check if supported
        if file_extension not in self.SUPPORTED_TYPES:
            # Try to detect MIME type
            mime_type, _ = mimetypes.guess_type(str(file_path))
            
            # Map MIME type to extension
            for ext, supported_mime in self.SUPPORTED_TYPES.items():
                if mime_type == supported_mime:
                    file_extension = ext
                    break
            else:
                raise create_error(
                    UnsupportedFileTypeError,
                    f"Unsupported file type: {file_extension or mime_type}",
                    "UNSUPPORTED_TYPE",
                    filename=filename,
                    detected_mime=mime_type
                )
        
        logger.debug("File type classified", 
                    filename=filename, 
                    type=file_extension)
        
        return file_extension.lstrip('.')
    
    async def _extract_text(self, file_path: Path, file_type: str) -> Tuple[str, str, Optional[int]]:
        """
        Extract text based on file type.
        
        Args:
            file_path: Path to the file
            file_type: File type (without dot)
            
        Returns:
            Tuple[str, str, Optional[int]]: Extracted text, extraction method, total pages
            
        Raises:
            DocumentExtractionError: If extraction fails
        """
        try:
            if file_type == 'pdf':
                return await self._extract_pdf_text(file_path)
            elif file_type == 'docx':
                return self._extract_docx_text(file_path)
            elif file_type == 'xlsx':
                return self._extract_xlsx_text(file_path)
            elif file_type == 'txt':
                return self._extract_txt_text(file_path)
            elif file_type == 'eml':
                return self._extract_eml_text(file_path)
            elif file_type in ['png', 'jpg', 'jpeg']:
                return self._extract_image_text(file_path)
            elif file_type == 'zip':
                return await self._extract_zip_text(file_path)
            else:
                raise create_error(
                    UnsupportedFileTypeError,
                    f"No extraction method for file type: {file_type}",
                    "NO_EXTRACTOR",
                    file_type=file_type
                )
                
        except Exception as e:
            if isinstance(e, (UnsupportedFileTypeError, DocumentExtractionError)):
                raise
            
            raise create_error(
                DocumentExtractionError,
                f"Text extraction failed: {str(e)}",
                "EXTRACTION_FAILED",
                file_type=file_type,
                error_type=type(e).__name__
            )
    
    async def _extract_pdf_text(self, file_path: Path) -> Tuple[str, str, int]:
        """
        Extract text from PDF using dual-strategy approach.
        
        Uses PyMuPDF4LLM and PDFPlumber in parallel, choosing the best result per page.
        """
        logger.debug("Starting PDF extraction", file=str(file_path))
        
        try:
            # Extract with PyMuPDF4LLM
            pymupdf_text = pymupdf4llm.to_markdown(str(file_path))
            
            # Extract with PDFPlumber for comparison
            with pdfplumber.open(file_path) as pdf:
                pdfplumber_pages = []
                for page in pdf.pages:
                    page_text = page.extract_text() or ""
                    pdfplumber_pages.append(page_text)
                
                total_pages = len(pdf.pages)
            
            # Split PyMuPDF text by pages (approximate)
            pymupdf_pages = self._split_text_by_pages(pymupdf_text, total_pages)
            
            # Choose best extraction per page (95% threshold)
            final_pages = []
            pymupdf_count = 0
            pdfplumber_count = 0
            
            for i in range(total_pages):
                pymupdf_page = pymupdf_pages[i] if i < len(pymupdf_pages) else ""
                pdfplumber_page = pdfplumber_pages[i] if i < len(pdfplumber_pages) else ""
                
                # Compare character counts
                if len(pymupdf_page) >= 0.95 * len(pdfplumber_page):
                    final_pages.append(pymupdf_page)
                    pymupdf_count += 1
                else:
                    final_pages.append(pdfplumber_page)
                    pdfplumber_count += 1
            
            final_text = "\n\n".join(final_pages)
            method = f"hybrid_pdf (pymupdf: {pymupdf_count}, pdfplumber: {pdfplumber_count})"
            
            logger.info("PDF extraction completed",
                       total_pages=total_pages,
                       text_length=len(final_text),
                       method=method)
            
            return final_text, method, total_pages
            
        except Exception as e:
            raise create_error(
                DocumentExtractionError,
                f"PDF extraction failed: {str(e)}",
                "PDF_EXTRACTION_FAILED"
            )
    
    def _extract_docx_text(self, file_path: Path) -> Tuple[str, str, None]:
        """Extract text from DOCX file."""
        try:
            doc = DocxDocument(file_path)
            paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
            text = "\n\n".join(paragraphs)
            
            logger.debug("DOCX extraction completed", 
                        paragraphs=len(paragraphs),
                        text_length=len(text))
            
            return text, "python-docx", None
            
        except Exception as e:
            raise create_error(
                DocumentExtractionError,
                f"DOCX extraction failed: {str(e)}",
                "DOCX_EXTRACTION_FAILED"
            )
    
    def _extract_xlsx_text(self, file_path: Path) -> Tuple[str, str, None]:
        """Extract text from Excel file."""
        try:
            workbook = load_workbook(file_path, data_only=True)
            all_text = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_text = [f"Sheet: {sheet_name}"]
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    if row_text.strip():
                        sheet_text.append(row_text)
                
                all_text.extend(sheet_text)
                all_text.append("")  # Separator between sheets
            
            text = "\n".join(all_text)
            
            logger.debug("XLSX extraction completed",
                        sheets=len(workbook.sheetnames),
                        text_length=len(text))
            
            return text, "openpyxl", None
            
        except Exception as e:
            raise create_error(
                DocumentExtractionError,
                f"XLSX extraction failed: {str(e)}",
                "XLSX_EXTRACTION_FAILED"
            )
    
    def _extract_txt_text(self, file_path: Path) -> Tuple[str, str, None]:
        """Extract text from plain text file."""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                text = f.read()
            
            logger.debug("TXT extraction completed", text_length=len(text))
            return text, "plain_text", None
            
        except Exception as e:
            raise create_error(
                DocumentExtractionError,
                f"TXT extraction failed: {str(e)}",
                "TXT_EXTRACTION_FAILED"
            )
    
    def _extract_eml_text(self, file_path: Path) -> Tuple[str, str, None]:
        """Extract text from email file."""
        try:
            with open(file_path, 'rb') as f:
                msg = email.message_from_bytes(f.read(), policy=default)
            
            # Extract headers
            headers = []
            for header in ['From', 'To', 'Subject', 'Date']:
                if msg[header]:
                    headers.append(f"{header}: {msg[header]}")
            
            # Extract body
            body_parts = []
            if msg.is_multipart():
                for part in msg.walk():
                    if part.get_content_type() == "text/plain":
                        body_parts.append(part.get_content())
            else:
                body_parts.append(msg.get_content())
            
            text = "\n".join(headers) + "\n\n" + "\n".join(body_parts)
            
            logger.debug("EML extraction completed", text_length=len(text))
            return text, "email_parser", None
            
        except Exception as e:
            raise create_error(
                DocumentExtractionError,
                f"EML extraction failed: {str(e)}",
                "EML_EXTRACTION_FAILED"
            )
    
    def _extract_image_text(self, file_path: Path) -> Tuple[str, str, None]:
        """Extract text from image using OCR."""
        try:
            # Check if Tesseract is available before attempting OCR
            try:
                pytesseract.get_tesseract_version()
            except pytesseract.TesseractNotFoundError:
                raise create_error(
                    DocumentExtractionError,
                    "Tesseract OCR not found. Please install Tesseract or set TESSERACT_CMD environment variable.",
                    "TESSERACT_NOT_FOUND"
                )
            
            image = Image.open(file_path)
            text = pytesseract.image_to_string(image)
            
            logger.debug("Image OCR completed", text_length=len(text))
            return text, "pytesseract_ocr", None
            
        except DocumentExtractionError:
            # Re-raise our custom errors
            raise
        except Exception as e:
            raise create_error(
                DocumentExtractionError,
                f"Image OCR failed: {str(e)}",
                "OCR_EXTRACTION_FAILED"
            )
    
    async def _extract_zip_text(self, file_path: Path) -> Tuple[str, str, None]:
        """Extract text from ZIP archive by processing contained files."""
        try:
            all_text = []
            extraction_methods = []
            
            with zipfile.ZipFile(file_path, 'r') as zip_file:
                for file_info in zip_file.infolist():
                    if not file_info.is_dir():
                        # Extract file to downloads directory
                        extracted_path = self.downloads_dir / file_info.filename
                        extracted_path.parent.mkdir(parents=True, exist_ok=True)
                        
                        with zip_file.open(file_info) as source, open(extracted_path, 'wb') as target:
                            target.write(source.read())
                        
                        try:
                            # Recursively process extracted file
                            file_type = self._classify_file_type(extracted_path, file_info.filename)
                            file_text, method, _ = await self._extract_text(extracted_path, file_type)
                            
                            all_text.append(f"File: {file_info.filename}\n{file_text}")
                            extraction_methods.append(method)
                            
                        except (UnsupportedFileTypeError, DocumentExtractionError):
                            # Skip unsupported files in ZIP
                            logger.warning("Skipping unsupported file in ZIP", 
                                         filename=file_info.filename)
                            continue
            
            text = "\n\n" + "="*50 + "\n\n".join(all_text)
            method = f"zip_recursive ({', '.join(set(extraction_methods))})"
            
            logger.debug("ZIP extraction completed",
                        files_processed=len(all_text),
                        text_length=len(text))
            
            return text, method, None
            
        except Exception as e:
            raise create_error(
                DocumentExtractionError,
                f"ZIP extraction failed: {str(e)}",
                "ZIP_EXTRACTION_FAILED"
            )
    
    def _split_text_by_pages(self, text: str, total_pages: int) -> List[str]:
        """Split text into approximate pages."""
        if total_pages <= 1:
            return [text]
        
        # Simple heuristic: split by length
        chars_per_page = len(text) // total_pages
        pages = []
        
        for i in range(total_pages):
            start = i * chars_per_page
            end = (i + 1) * chars_per_page if i < total_pages - 1 else len(text)
            pages.append(text[start:end])
        
        return pages
    
    def _generate_content_hash(self, content: str) -> str:
        """Generate SHA-256 hash of content for deduplication."""
        return hashlib.sha256(content.encode('utf-8')).hexdigest()
    
    def _extract_first_pages(self, text: str, num_pages: int) -> str:
        """Extract first N pages worth of content."""
        # Simple heuristic: take first 20% of text or 2000 characters per "page"
        chars_per_page = 2000
        max_chars = chars_per_page * num_pages
        
        return text[:max_chars] if len(text) > max_chars else text
    
    def _cleanup_temp_files(self, file_path: Optional[Path]) -> None:
        """
        Note: Files are no longer cleaned up as they are stored permanently 
        in the downloads directory for data management purposes.
        """
        # Files are now stored permanently in downloads/{timestamp} directories
        # This method is kept for backward compatibility but does nothing
        pass
